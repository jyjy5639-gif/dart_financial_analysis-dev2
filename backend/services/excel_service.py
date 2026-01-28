from io import BytesIO
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.drawing.image import Image as OpenpyxlImage
import base64
from backend.services.chart_service import ChartService


class ExcelService:
    """재무 데이터를 엑셀로 변환하는 서비스 (차트 포함)"""
    
    @staticmethod
    def create_financial_excel(financial_data: List[Dict[str, Any]]) -> BytesIO:
        """
        재무 데이터를 엑셀 파일로 변환 (차트 포함)
        
        Args:
            financial_data: 재무 데이터 리스트 (각 회사별 데이터)
            
        Returns:
            BytesIO: 엑셀 파일 바이너리
        """
        wb = Workbook()
        wb.remove(wb.active)  # 기본 시트 제거
        
        # 차트 서비스 인스턴스
        chart_service = ChartService()
        
        companies_summary = []
        
        # 복수 회사인 경우 요약 비교 시트를 먼저 생성
        if len(financial_data) > 1:
            companies_summary = financial_data.copy()
            ExcelService._create_summary_sheet(wb, companies_summary, chart_service)
        
        # 각 회사별 시트 생성
        for company_data in financial_data:
            company_name = company_data.get('company_name', '알 수 없음')
            stock_code = company_data.get('stock_code', '')
            financial_statements = company_data.get('financial_statements', {})
            
            # 회사별 재무제표 시트 생성
            sheet_title = f"{company_name}"
            if stock_code and stock_code != 'N/A':
                sheet_title += f"_{stock_code}"
            sheet_title = sheet_title[:31]  # 시트명 길이 제한
            
            ws = wb.create_sheet(title=sheet_title)
            
            ExcelService._add_company_header(ws, company_name, stock_code)
            
            row_offset = 3
            
            # 차트 추가 (단일 회사)
            row_offset = ExcelService._add_single_company_charts(
                ws, company_data, chart_service, row_offset
            )
            row_offset += 2
            
            # 재무상태표
            if 'balance_sheet' in financial_statements and financial_statements['balance_sheet']:
                row_offset = ExcelService._add_financial_table(
                    ws, "재무상태표", financial_statements['balance_sheet'], row_offset
                )
                row_offset += 2
            
            # 손익계산서
            if 'income_statement' in financial_statements and financial_statements['income_statement']:
                row_offset = ExcelService._add_financial_table(
                    ws, "손익계산서", financial_statements['income_statement'], row_offset
                )
                row_offset += 2
            
            # 현금흐름표
            if 'cash_flow' in financial_statements and financial_statements['cash_flow']:
                row_offset = ExcelService._add_financial_table(
                    ws, "현금흐름표", financial_statements['cash_flow'], row_offset
                )
                row_offset += 2
            
            # 재무 비율
            if 'ratios' in financial_statements and financial_statements['ratios']:
                row_offset = ExcelService._add_ratios_table(
                    ws, financial_statements['ratios'], row_offset
                )
        
        # 엑셀 파일을 BytesIO로 저장
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    @staticmethod
    def _add_single_company_charts(ws, company_data: Dict, chart_service: ChartService, start_row: int) -> int:
        """단일 회사 차트 추가"""
        try:
            # 재무 데이터를 차트 서비스 형식으로 변환
            financial_items = ExcelService._convert_to_chart_format(company_data)
            ratios = company_data.get('financial_statements', {}).get('ratios', {})
            
            # 트렌드 차트 생성
            trend_image = chart_service.create_trend_chart(
                financial_items, 
                chart_type="matplotlib"
            )
            
            if trend_image:
                ExcelService._add_image_to_sheet(ws, trend_image, start_row, 1)
                start_row += 20
            
            # 비율 차트 생성
            if ratios:
                ratio_image = chart_service.create_ratio_chart(
                    ratios,
                    chart_type="matplotlib"
                )
                
                if ratio_image:
                    ExcelService._add_image_to_sheet(ws, ratio_image, start_row, 1)
                    start_row += 25
            
        except Exception as e:
            print(f"차트 추가 실패: {e}")
        
        return start_row
    
    @staticmethod
    def _add_comparison_charts(ws, companies_summary: List[Dict], chart_service: ChartService, start_row: int) -> int:
        """비교 차트 추가"""
        try:
            # 회사별 데이터를 차트 서비스 형식으로 변환
            companies_data = {}
            
            for idx, company_data in enumerate(companies_summary):
                company_name = company_data.get('company_name', f'회사{idx+1}')
                
                # financial_data 형식으로 변환
                financial_items = ExcelService._convert_to_chart_format(company_data)
                ratios = company_data.get('financial_statements', {}).get('ratios', {})
                
                companies_data[f'corp_{idx}'] = {
                    'corp_name': company_name,
                    'financial_data': financial_items,
                    'ratios': ratios
                }
            
            # 비교 차트 생성
            comparison_image = chart_service.create_comparison_chart(
                companies_data,
                chart_type="matplotlib"
            )
            
            if comparison_image:
                ExcelService._add_image_to_sheet(ws, comparison_image, start_row, 1)
                start_row += 20
            
            # 비율 비교 차트 생성
            ratio_comparison_image = chart_service.create_ratio_comparison_chart(
                companies_data,
                chart_type="matplotlib"
            )
            
            if ratio_comparison_image:
                ExcelService._add_image_to_sheet(ws, ratio_comparison_image, start_row, 1)
                start_row += 25
            
        except Exception as e:
            print(f"비교 차트 추가 실패: {e}")
        
        return start_row
    
    @staticmethod
    def _convert_to_chart_format(company_data: Dict) -> List[Dict]:
        """회사 데이터를 차트 서비스가 기대하는 형식으로 변환"""
        financial_items = []
        financial_statements = company_data.get('financial_statements', {})
        
        # 주요 계정들 매핑
        account_mapping = {
            '매출액': ['매출액', '수익(매출액)'],
            '영업이익': ['영업이익', '영업이익(손실)'],
            '당기순이익': ['당기순이익', '당기순이익(손실)'],
            '자산총계': ['자산총계'],
            '부채총계': ['부채총계'],
            '자본총계': ['자본총계']
        }
        
        for standard_name, possible_names in account_mapping.items():
            # 손익계산서와 재무상태표에서 찾기
            found_data = None
            
            for statement_type in ['income_statement', 'balance_sheet']:
                statement_data = financial_statements.get(statement_type, {})
                
                for possible_name in possible_names:
                    if possible_name in statement_data:
                        found_data = statement_data[possible_name]
                        break
                
                if found_data:
                    break
            
            if found_data:
                # 연도별 데이터를 추출하여 차트 형식으로 변환
                years = sorted(found_data.keys(), reverse=False)  # 오래된 순으로
                
                item = {
                    'base_display_name': standard_name,
                    'display_name': standard_name
                }
                
                # 최대 3년치 데이터
                if len(years) >= 3:
                    item['bfefrmtrm_amount'] = found_data.get(years[0], 0)
                    item['bfefrmtrm_dt'] = f"{years[0]}1231"
                    item['frmtrm_amount'] = found_data.get(years[1], 0) 
                    item['frmtrm_dt'] = f"{years[1]}1231"
                    item['thstrm_amount'] = found_data.get(years[2], 0)
                    item['thstrm_dt'] = f"{years[2]}1231"
                elif len(years) >= 2:
                    item['frmtrm_amount'] = found_data.get(years[0], 0)
                    item['frmtrm_dt'] = f"{years[0]}1231"
                    item['thstrm_amount'] = found_data.get(years[1], 0)
                    item['thstrm_dt'] = f"{years[1]}1231"
                elif len(years) >= 1:
                    item['thstrm_amount'] = found_data.get(years[0], 0)
                    item['thstrm_dt'] = f"{years[0]}1231"
                
                financial_items.append(item)
        
        return financial_items
    
    @staticmethod
    def _add_image_to_sheet(ws, image_base64: str, start_row: int, start_col: int):
        """base64 인코딩된 이미지를 시트에 추가"""
        try:
            # base64 디코딩
            image_data = base64.b64decode(image_base64)
            
            # BytesIO를 사용하여 메모리에서 직접 이미지 처리
            image_stream = BytesIO(image_data)
            
            # 이미지를 워크시트에 추가
            img = OpenpyxlImage(image_stream)
            img.width = 600  # 너비 조정
            img.height = 400  # 높이 조정
            
            # 셀 위치 계산 (A1 스타일)
            cell_address = f"{chr(64 + start_col)}{start_row}"
            img.anchor = cell_address
            
            ws.add_image(img)
                    
        except Exception as e:
            print(f"이미지 추가 실패: {e}")
    
    @staticmethod
    def _add_company_header(ws, company_name: str, stock_code: str):
        """회사 정보 헤더 추가"""
        header_text = f"{company_name}"
        if stock_code and stock_code != 'N/A':
            header_text += f" ({stock_code})"
        header_text += " 재무제표"
        
        ws['A1'] = header_text
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
    
    @staticmethod
    def _add_financial_table(ws, table_name: str, data: Dict, start_row: int) -> int:
        """재무제표 테이블 추가"""
        # 테이블 제목
        ws.cell(row=start_row, column=1, value=table_name)
        ws.cell(row=start_row, column=1).font = Font(size=14, bold=True)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
        
        current_row = start_row + 1
        
        if not data:
            ws.cell(row=current_row, column=1, value="데이터 없음")
            return current_row + 1
        
        # 연도 헤더 추출 (최근 3년)
        years = sorted(set(year for item in data.values() for year in item.keys() if year), reverse=True)[:3]
        
        if not years:
            ws.cell(row=current_row, column=1, value="데이터 없음")
            return current_row + 1
        
        # 헤더 행
        ws.cell(row=current_row, column=1, value="항목")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        for idx, year in enumerate(years, start=2):
            ws.cell(row=current_row, column=idx, value=f"{year}년")
            ws.cell(row=current_row, column=idx).font = Font(bold=True)
            ws.cell(row=current_row, column=idx).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            ws.cell(row=current_row, column=idx).alignment = Alignment(horizontal='right')
        
        current_row += 1
        
        # 데이터 행
        for item_name, item_data in data.items():
            ws.cell(row=current_row, column=1, value=item_name)
            
            for idx, year in enumerate(years, start=2):
                value = item_data.get(year, '')
                
                # 숫자 변환 시도
                try:
                    if isinstance(value, str):
                        value = value.replace(',', '')
                    numeric_value = float(value) if value else 0
                    ws.cell(row=current_row, column=idx, value=numeric_value)
                    ws.cell(row=current_row, column=idx).number_format = '#,##0'
                except (ValueError, TypeError):
                    ws.cell(row=current_row, column=idx, value=value)
                
                ws.cell(row=current_row, column=idx).alignment = Alignment(horizontal='right')
            
            current_row += 1
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 35
        for idx in range(2, len(years) + 2):
            ws.column_dimensions[chr(64 + idx)].width = 18
        
        return current_row
    
    @staticmethod
    def _add_ratios_table(ws, ratios: Dict, start_row: int) -> int:
        """재무 비율 테이블 추가"""
        # 테이블 제목
        ws.cell(row=start_row, column=1, value="재무 비율")
        ws.cell(row=start_row, column=1).font = Font(size=14, bold=True)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
        
        current_row = start_row + 1
        
        if not ratios:
            ws.cell(row=current_row, column=1, value="데이터 없음")
            return current_row + 1
        
        # 헤더 행
        headers = ["비율명", "당기", "전기", "전전기"]
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=current_row, column=idx, value=header)
            ws.cell(row=current_row, column=idx).font = Font(bold=True)
            ws.cell(row=current_row, column=idx).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            if idx > 1:
                ws.cell(row=current_row, column=idx).alignment = Alignment(horizontal='right')
        
        current_row += 1
        
        # 데이터 행
        for ratio_name, ratio_values in ratios.items():
            ws.cell(row=current_row, column=1, value=ratio_name)
            
            # 당기, 전기, 전전기
            for idx, period in enumerate(['thstrm', 'frmtrm', 'bfefrmtrm'], start=2):
                value = ratio_values.get(period, 0)
                ws.cell(row=current_row, column=idx, value=value)
                ws.cell(row=current_row, column=idx).number_format = '0.00"%"'
                ws.cell(row=current_row, column=idx).alignment = Alignment(horizontal='right')
            
            current_row += 1
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D']:
            ws.column_dimensions[col].width = 15
        
        return current_row
    
    @staticmethod
    def _create_summary_sheet(wb, companies_summary: List[Dict], chart_service: ChartService):
        """회사 간 비교 요약 시트 생성 (차트 포함)"""
        ws = wb.create_sheet(title="📊 요약_비교", index=0)
        
        ws['A1'] = "회사별 재무 비교 요약"
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 병합할 열 개수 계산 (회사 수 + 1)
        merge_end_col = len(companies_summary) + 1
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_end_col)
        ws.row_dimensions[1].height = 30
        
        current_row = 3
        
        # 비교 차트 추가
        current_row = ExcelService._add_comparison_charts(
            ws, companies_summary, chart_service, current_row
        )
        current_row += 2
        
        # 주요 지표 정의 (손익계산서)
        income_metrics = [
            ('매출액', 'income_statement', ['매출액', '수익(매출액)']),
            ('영업이익', 'income_statement', ['영업이익', '영업이익(손실)']),
            ('당기순이익', 'income_statement', ['당기순이익', '당기순이익(손실)']),
        ]
        
        # 재무상태표 지표
        balance_metrics = [
            ('총자산', 'balance_sheet', ['자산총계']),
            ('총부채', 'balance_sheet', ['부채총계']),
            ('자본총계', 'balance_sheet', ['자본총계']),
        ]
        
        # 재무비율
        ratio_metrics = [
            ('영업이익률', 'ratios'),
            ('순이익률', 'ratios'),
            ('ROE', 'ratios'),
            ('ROA', 'ratios'),
            ('부채비율', 'ratios'),
        ]
        
        # 섹션 1: 손익계산서 지표
        current_row = ExcelService._add_summary_section(
            ws, "손익계산서 주요 지표", income_metrics, companies_summary, current_row
        )
        current_row += 2
        
        # 섹션 2: 재무상태표 지표
        current_row = ExcelService._add_summary_section(
            ws, "재무상태표 주요 지표", balance_metrics, companies_summary, current_row
        )
        current_row += 2
        
        # 섹션 3: 재무비율
        current_row = ExcelService._add_ratio_summary_section(
            ws, "재무 비율", ratio_metrics, companies_summary, current_row
        )
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 20
        for idx in range(len(companies_summary)):
            ws.column_dimensions[chr(66 + idx)].width = 20
    
    @staticmethod
    def _add_summary_section(ws, section_name: str, metrics: List, companies_summary: List[Dict], start_row: int) -> int:
        """요약 섹션 추가 (재무제표 항목)"""
        current_row = start_row
        
        # 섹션 제목
        ws.cell(row=current_row, column=1, value=section_name)
        ws.cell(row=current_row, column=1).font = Font(size=14, bold=True)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(companies_summary) + 1)
        current_row += 1
        
        # 테이블 헤더
        ws.cell(row=current_row, column=1, value="항목")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        for idx, company in enumerate(companies_summary, start=2):
            company_name = company['company_name']
            stock_code = company.get('stock_code', '')
            cell_value = f"{company_name}"
            if stock_code and stock_code != 'N/A':
                cell_value += f"\n({stock_code})"
            
            ws.cell(row=current_row, column=idx, value=cell_value)
            ws.cell(row=current_row, column=idx).font = Font(bold=True)
            ws.cell(row=current_row, column=idx).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws.cell(row=current_row, column=idx).alignment = Alignment(horizontal='center', wrap_text=True)
        
        current_row += 1
        
        # 각 지표별 데이터
        for metric_name, statement_type, item_names in metrics:
            ws.cell(row=current_row, column=1, value=metric_name)
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            
            for idx, company in enumerate(companies_summary, start=2):
                statements = company['financial_statements'].get(statement_type, {})
                
                # 여러 가능한 항목명 중 첫 번째로 찾은 것 사용
                value = None
                for item_name in item_names:
                    if item_name in statements:
                        data = statements[item_name]
                        # 최신 연도 데이터 추출
                        if data:
                            years = sorted(data.keys(), reverse=True)
                            if years:
                                value = data[years[0]]
                                break
                
                # 셀에 값 입력
                try:
                    if value:
                        if isinstance(value, str):
                            value = value.replace(',', '')
                        numeric_value = float(value)
                        ws.cell(row=current_row, column=idx, value=numeric_value)
                        ws.cell(row=current_row, column=idx).number_format = '#,##0'
                    else:
                        ws.cell(row=current_row, column=idx, value='-')
                except (ValueError, TypeError):
                    ws.cell(row=current_row, column=idx, value=value if value else '-')
                
                ws.cell(row=current_row, column=idx).alignment = Alignment(horizontal='right')
            
            current_row += 1
        
        return current_row
    
    @staticmethod
    def _add_ratio_summary_section(ws, section_name: str, metrics: List, companies_summary: List[Dict], start_row: int) -> int:
        """요약 섹션 추가 (재무비율)"""
        current_row = start_row
        
        # 섹션 제목
        ws.cell(row=current_row, column=1, value=section_name)
        ws.cell(row=current_row, column=1).font = Font(size=14, bold=True)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(companies_summary) + 1)
        current_row += 1
        
        # 테이블 헤더
        ws.cell(row=current_row, column=1, value="비율명")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        for idx, company in enumerate(companies_summary, start=2):
            company_name = company['company_name']
            stock_code = company.get('stock_code', '')
            cell_value = f"{company_name}"
            if stock_code and stock_code != 'N/A':
                cell_value += f"\n({stock_code})"
            
            ws.cell(row=current_row, column=idx, value=cell_value)
            ws.cell(row=current_row, column=idx).font = Font(bold=True)
            ws.cell(row=current_row, column=idx).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws.cell(row=current_row, column=idx).alignment = Alignment(horizontal='center', wrap_text=True)
        
        current_row += 1
        
        # 각 비율별 데이터
        for ratio_name, _ in metrics:
            ws.cell(row=current_row, column=1, value=ratio_name)
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            
            for idx, company in enumerate(companies_summary, start=2):
                ratios = company['financial_statements'].get('ratios', {})
                
                if ratio_name in ratios:
                    value = ratios[ratio_name].get('thstrm', 0)
                    ws.cell(row=current_row, column=idx, value=value)
                    ws.cell(row=current_row, column=idx).number_format = '0.00"%"'
                else:
                    ws.cell(row=current_row, column=idx, value='-')
                
                ws.cell(row=current_row, column=idx).alignment = Alignment(horizontal='right')
            
            current_row += 1
        
        return current_row
